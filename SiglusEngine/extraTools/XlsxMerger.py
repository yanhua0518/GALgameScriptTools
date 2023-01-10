# -*- coding: utf-8 -*-
# For Windows OS Only....

import sys
import os
import glob
from copy import copy,deepcopy
import openpyxl

def main(argv):
    if argv.count("-t"):
        addFN=True
        argv.remove("-t")
    else:
        addFN=False
    if len(argv)<2 or argv[1]=='':
        print ("Usage: "+argv[0][argv[0].rfind("/")+1:]+" <Excel folder\> [-t]")
        return False

    inF=argv[1]+"/"
    outXLS=argv[0][:argv[0].rfind("/")+1]+argv[1]+".xlsx"
    mixBook=openpyxl.Workbook()
    mixBook.remove(mixBook.active)
    for inFN in glob.glob(inF+"*.xlsx"):
        print(inFN)
        workBook=openpyxl.load_workbook(inFN)
        for sheet in workBook:
            title=inFN[inFN.rfind("\\")+1:].replace("[","(").replace("]",")").replace(".xlsx","")
            #print(title)
            if sheet.max_column==1 and sheet.max_row==1:
                continue
            if addFN:
                workSheet=mixBook.create_sheet(title+" "+sheet.title)
            else:
                workSheet=mixBook.create_sheet(sheet.title)
            print(workSheet.title)
            for row in sheet.rows:
                for cell in row:
                    workSheet[cell.coordinate].value=copy(cell.value)
                    workSheet[cell.coordinate].font=copy(cell.font)
                    workSheet[cell.coordinate].border=copy(cell.border)
                    workSheet[cell.coordinate].fill=copy(cell.fill)
                    workSheet[cell.coordinate].number_format=copy(cell.number_format)
                    workSheet[cell.coordinate].alignment=copy(cell.alignment)
            for i in sheet.column_dimensions:
                workSheet.column_dimensions[i].width=sheet.column_dimensions[i].width
            for i in sheet.row_dimensions:
                workSheet.row_dimensions[i].height=sheet.row_dimensions[i].height
                

    mixBook.save(outXLS)
    return True

if __name__=="__main__":
    main(sys.argv)
