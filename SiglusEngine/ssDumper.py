# -*- coding: utf-8 -*-
# For Windows OS Only....

import sys
import os
import glob
import struct
import unicodedata
from Decryption import Decrypt


class Header:
    headerData=b''
    headerList=[]
    length=0
    index=0
    count=0
    offset=0
    dataCount=0
    def __init__(H,f):
        f.seek(0)
        H.length=struct.unpack('I',f.read(4))[0]
        H.headerData=f.read(128)
        H.headerList=struct.unpack('32I',H.headerData)
        H.index=H.headerList[2]
        H.count=H.headerList[3]
        H.offset=H.headerList[4]
        H.dataCount=H.headerList[5]


def check(scr,checkHalf):
    if checkHalf:
        for char in scr:
            if unicodedata.east_asian_width(char)=='Na':
                return False
        return True
    else:
        for char in scr:
            if unicodedata.east_asian_width(char)!='Na':
                return True
        return False

def softIndex(l,c,e):
    return l*1000000+c*1000+e

def main(argv):
    if argv.count('-a')>0:
        noDump=True
        argv.remove('-a')
    else:
        noDump=False
    if argv.count('-w')>0:
        fullDump=True
        argv.remove('-w')
    else:
        fullDump=False
    if argv.count('-d')>0:
        copyLine=True
        argv.remove('-d')
    else:
        copyLine=False
    if argv.count('-x')>0:
        xlsxMode=True
        import openpyxl
        argv.remove('-x')
    else:
        xlsxMode=False
    if argv.count('-s')>0:
        singleXlsx=True
        argv.remove('-s')
    else:
        singleXlsx=False
    if argv.count('-c')>0:
        countWords=True
        argv.remove('-c')
    else:
        countWords=False
    if argv.count('-o')>0:
        offMode=True
        argv.remove('-o')
        '''
        LINE=1000000
        CHAR=1000
        END=1
        '''
    else:
        offMode=False
    
    if len(argv)<2 or argv[1]=='':
        print ("Usage: "+argv[0][argv[0].rfind("\\")+1:]+" <Scene\> [Text\] [-o] [-d] [-a/-w] [-x [-s [-c]]]")
        return False
    countError=0
    inF=argv[1]+"\\"
    if len(argv)<3 or argv[2]=='':
        outF=argv[0][:argv[0].rfind("\\")+1]+argv[1][argv[1].rfind("\\")+1:]+"_out\\"
    else:
        outF=argv[0][:argv[0].rfind("\\")+1]+argv[2]+"\\"

    if not singleXlsx and not os.path.exists(outF):
        os.makedirs(outF)
    if xlsxMode and singleXlsx:
        workBook=openpyxl.Workbook()
        if countWords:
            countSheet=workBook.active
            countSheet.title='Statistics'
            countSheet.column_dimensions['A'].width=40
        outXLS=outF[:outF.rfind("\\")].replace(".xlsx","")+".xlsx"
    
    for inFN in glob.glob(inF+"*.ss"):
        print(inFN)
        if xlsxMode:
            name=inFN[inFN.rfind("\\")+1:]
            sheet=name[:31]
            countLen=0
            if not singleXlsx:
                outXLS=outF+inFN[inFN.rfind("\\")+1:]+".xlsx"
                workBook=openpyxl.Workbook()
                workSheet=workBook.active
                workSheet.title=sheet
            else:
                workSheet=workBook.create_sheet(sheet)
            workSheet.column_dimensions['A'].width=6
            workSheet.column_dimensions['B'].width=64
            workSheet.column_dimensions['C'].width=64
            fillColor=openpyxl.styles.PatternFill(fill_type="solid", fgColor="F2F2F2")
            border=openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin',color='D0D7E5'),left=openpyxl.styles.Side(style='thin',color='D0D7E5'),right=openpyxl.styles.Side(style='thin',color='D0D7E5'))
            if name!=sheet:
                tempName=name
            else:
                tempName=""
            workSheet.append(["Index","Text","Translation",tempName])
        else:
            outFN=outF+inFN[inFN.rfind("\\")+1:]+".txt"
            try:
                output=open(outFN,'w',1,"UTF-8")
            except:
                print("Output file error!")
                countError+=1
                continue
        indexs=[]
        texts=[]
        if offMode:
            try:
                file=open(inFN,'r',1,"UTF-8")
                lines=file.readlines()
                file.close()
            except UnicodeDecodeError:
                try:
                    file=open(inFN,'r',1,"SHIFT-JIS")
                    lines=file.readlines()
                    file.close()
                except:
                    print("Input file error!")
                    countError+=1
                    continue
            except:
                print("Input file error!")
                countError+=1
                continue
            isNote=False
            for i,line in enumerate(lines):
                isText=0
                isAt=False
                start=0
                for n,char in enumerate(line):
                    if char=='/':
                        if isNote:
                            if line[n-1:n]=='*':
                                isNote=False
                        else:
                            if line[n+1:n+2]=='*':
                                isNote=True
                            elif line[n+1:n+2]=='/' and isText!=2:
                                break
                    if isNote or (isText==2 and char!='"'):
                        continue
                    if char=='#' or char==';':
                        break
                    if char=='\t' or char=='\n':
                        isAt=False
                        continue
                    if char=='@' and line[n+1:n+5]!='ruby':
                        isAt=True
                    if isAt:
                        if char==' ' or char==',' or char=='(' or char==')' or char=='【' or char=='「':
                            isAt=False
                    if char=='"':
                        if isText==2:
                            isText=0
                            if line[start:n]!='':
                                indexs.append(softIndex(i,start,n-1))
                                texts.append(line[start:n])
                        else:
                            if isText==1:
                                indexs.append(softIndex(i,start,n-1))
                                texts.append(line[start:n])
                            isText=2
                            start=n+1
                    elif unicodedata.east_asian_width(char)!='Na':
                        if not isText and not isAt:
                            isText=1
                            start=n
                    elif isText==1:
                        isText=0
                        indexs.append(softIndex(i,start,n-1))
                        texts.append(line[start:n])
                if isText:
                    isText=0
                    indexs.append(softIndex(i,start,n))
                    texts.append(line[start:-1])   
        else:
            try:
                file=open(inFN,'rb')
                header=Header(file)
                file.seek(header.index)
                offset=[]
                length=[]
                for n in range(0,header.count):
                    offset.append(struct.unpack('I',file.read(4))[0])
                    length.append(struct.unpack('I',file.read(4))[0])
                indexs=[]
                texts=[]
                for x in range(0,header.count):
                    if length[x]==0:
                        continue
                    file.seek(header.offset+offset[x]*2,0)
                    string=file.read(length[x]*2)
                    text=Decrypt(string,length[x],x).decode("UTF-16")
                    indexs.append(x)
                    texts.append(text)
                file.close()
            except:
                print("Input file error!")
                countError+=1
        if indexs and texts:
            for index,text in zip(indexs,texts):
                if not check(text,fullDump) and not noDump:
                    continue
                if xlsxMode:
                    countLen+=len(text)
                    if copyLine:
                        workSheet.append([index,text,text])
                    else:
                        workSheet.append([index,text])
                    workSheet.cell(workSheet.max_row,3).fill=fillColor
                    workSheet.cell(workSheet.max_row,3).border=border
                else:
                    if copyLine:
                        outLine="○"+'%.10d'%index+"○"+text+"\n●"+'%.10d'%index+"●"+text+"\n\n"
                    else:
                        outLine="○"+'%.10d'%index+"○"+text+"\n●"+'%.10d'%index+"●\n\n"
                    output.write(outLine)
        if xlsxMode:
            textLine=workSheet.max_row
            if textLine==1 and singleXlsx:
                workBook.remove(workSheet)
            elif textLine>1 and not singleXlsx:
                try:
                    workBook.save(outXLS)
                except:
                    print("Output file error!")
                    countError+=1
            elif singleXlsx and countWords:
                countSheet.append([sheet,countLen])
        else:
            output.close()
            if os.path.getsize(outFN)==0:
                os.remove(outFN)
    if singleXlsx:
        print("Saving xlsx file...")
        if not countWords:
            workBook.remove(workBook['Sheet'])
        try:
            workBook.save(outXLS)
        except Exception as e:
            print("Output file error!")
            return e
            
    if countError:
        return countError
    else:
        return True

if __name__=="__main__":
    main(sys.argv)
